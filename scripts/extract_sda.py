#!/usr/bin/env python3
"""Extract the NDIA quarterly SDA supplement (Supplement P) into tidy JSON.

The published workbook is Strict OOXML, which openpyxl cannot open, and every
table is a wide human-readable cross-tab. This module normalises both problems
and derives the one measure the NDIA does not publish directly: SDA *places*
(resident capacity) by design category and region, which is the only unit that
is comparable with participant demand.

Usage:  python3 scripts/extract_sda.py <workbook.xlsx> [-o data/]
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl

# Strict OOXML uses purl.oclc.org namespaces; openpyxl only understands the
# transitional schemas. Rewriting the namespace URIs is a lossless fix.
STRICT_TO_TRANSITIONAL = [
    (b"http://purl.oclc.org/ooxml/spreadsheetml/main",
     b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"),
    (b"http://purl.oclc.org/ooxml/officeDocument/relationships",
     b"http://schemas.openxmlformats.org/officeDocument/2006/relationships"),
    (b"http://purl.oclc.org/ooxml/drawingml/main",
     b"http://schemas.openxmlformats.org/drawingml/2006/main"),
    (b"http://purl.oclc.org/ooxml/drawingml/chart",
     b"http://schemas.openxmlformats.org/drawingml/2006/chart"),
    (b"http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing",
     b"http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"),
    (b"http://purl.oclc.org/ooxml/officeDocument/extendedProperties",
     b"http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"),
    (b"http://purl.oclc.org/ooxml/officeDocument/docPropsVTypes",
     b"http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"),
    (b'conformance="strict"', b""),
]

DESIGN_CATEGORIES = [
    "Basic",
    "Improved Liveability",
    "High Physical Support",
    "Robust",
    "Fully Accessible",
    "Multi-Design Category",
]

# Design categories a participant can actually be found eligible for. "Basic" is
# no longer issued as an eligibility decision (folded into "Missing"), and
# "Multi-Design Category" is a dwelling attribute, not a decision, so neither
# has a demand-side counterpart to compare against.
COMPARABLE_CATEGORIES = [
    "Improved Liveability",
    "High Physical Support",
    "Robust",
    "Fully Accessible",
]

# Cells the NDIA suppresses for privacy, e.g. "<11", "<5", "n/a".
SUPPRESSED = re.compile(r"^\s*<\s*(\d+)\s*$")

# "Apartment, 2 bedrooms, 1 resident - Fully Accessible" and friends. The
# resident count is what turns a dwelling count into a places count; where a
# label carries both a bedroom and a resident figure, residents is the second.
DWELLING_LABEL = re.compile(
    r"^(?P<form>.*?),\s*(?P<first>\d+)\+?\s*(?:bedrooms?|residents?)"
    r"(?:,\s*(?P<residents>\d+)\s*residents?)?\s*-\s*(?P<category>.*)$"
)


def to_transitional(src: Path, dst: Path) -> Path:
    """Rewrite a Strict OOXML workbook into the transitional schema."""
    with zipfile.ZipFile(src) as zin, \
         zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith((".xml", ".rels")):
                for old, new in STRICT_TO_TRANSITIONAL:
                    data = data.replace(old, new)
            zout.writestr(item, data)
    return dst


def parse_value(raw):
    """Return (value, flag). Suppressed and unavailable cells keep their meaning.

    A suppressed cell is not zero and must never be summed as though it were, so
    it comes back as None alongside the ceiling the NDIA disclosed.
    """
    if raw is None:
        return None, "missing"
    if isinstance(raw, (int, float)):
        return float(raw), None
    text = str(raw).strip()
    match = SUPPRESSED.match(text)
    if match:
        return None, f"suppressed:<{match.group(1)}"
    if text.lower() in {"n/a", "na", "-", ""}:
        return None, "unavailable"
    cleaned = re.sub(r"[,$%\s]", "", text)
    try:
        return float(cleaned), None
    except ValueError:
        return None, "text"


def read_table(workbook, sheet_name):
    """Read one Supplement P sheet into {'columns': [...], 'rows': {label: {...}}}.

    Row 1 is the table caption, row 2 the header, then data rows followed by
    footnotes. Footnotes are long prose in the label column, so length
    discriminates them from region names.
    """
    sheet = workbook[sheet_name]
    raw_rows = list(sheet.iter_rows(values_only=True))
    header = [h for h in raw_rows[1] if h is not None]
    columns = [str(c).strip() for c in header[1:]]

    rows, notes = {}, []
    for raw in raw_rows[2:]:
        label = raw[0]
        if label is None:
            continue
        label = str(label).strip()
        if label.startswith("Back to") or len(label) > 60:
            if len(label) > 60:
                notes.append(label)
            continue
        record, flags = {}, {}
        for column, cell in zip(columns, raw[1:len(header)]):
            value, flag = parse_value(cell)
            record[column] = value
            if flag:
                flags[column] = flag
        rows[label] = {"values": record, "flags": flags}

    return {
        "caption": str(raw_rows[0][0]).strip() if raw_rows[0][0] else sheet_name,
        "columns": columns,
        "rows": rows,
        "notes": notes,
    }


def places_by_category(row_values):
    """Collapse a dwelling-type cross-tab row into places per design category.

    Each column label names a dwelling form and its resident count, so
    dwellings x residents gives resident capacity ("places") — the unit that is
    directly comparable with a participant count.
    """
    totals: dict[str, float] = {}
    for label, value in row_values.items():
        if label == "Total" or value is None:
            continue
        match = DWELLING_LABEL.match(label)
        if not match:
            continue
        residents = int(match.group("residents") or match.group("first"))
        category = match.group("category").strip()
        totals[category] = totals.get(category, 0.0) + value * residents
    return totals


def extract_national_trend(source: Path, figure_prose=()):
    """Recover the quarterly national series from Figure P.1.

    The figure's numbers are not in any cell: most live in the embedded chart
    XML, and the enrolled-dwellings series survives only in the figure's
    accessibility description. This is the workbook's only time series.
    """
    wanted = {
        "Participants with SDA in use": "sda_in_use",
        "Participants SDA eligible, not yet using SDA": "sda_eligible_not_using",
        "Active participants with SIL supports": "sil_participants",
        "Annualised SDA supports in active plans ($m)": "sda_annualised",
        "Annualised committed support for participants with SIL ($m)": "sil_annualised",
    }
    series, quarters = {}, []
    with zipfile.ZipFile(source) as archive:
        charts = [n for n in archive.namelist()
                  if re.match(r"xl/charts/chart\d+\.xml$", n)]
        for name in sorted(charts):
            xml = archive.read(name).decode("utf-8", "replace")
            # Excel parks filtered-out series in the c15 extension namespace, so
            # both tags must be split on or later series merge into one block.
            for block in re.split(r"<c(?:15)?:ser>", xml)[1:]:
                values = re.findall(r"<c:v>([^<]*)</c:v>", block)
                if not values:
                    continue
                label = values[0]
                periods = [v for v in values[1:]
                           if re.match(r"^[A-Z][a-z]{2}-\d{2}$", v)]
                numbers = [float(v) for v in values[1:]
                           if re.match(r"^-?\d+(\.\d+)?$", v)]
                if label in wanted and periods and len(periods) == len(numbers):
                    quarters = quarters or periods
                    series[wanted[label]] = numbers

    if "enrolled_dwellings" not in series:
        months = {m: i for i, m in enumerate(
            "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split(), 1)}
        found = {}
        for prose in figure_prose:
            if "enrolled dwellings" not in prose.lower():
                continue
            for name, year, count in re.findall(
                    r"In (\w+) (\d{4}) there were ([\d,]+) enrolled dwellings", prose):
                if name[:3] in months:
                    found[f"{name[:3]}-{year[2:]}"] = float(count.replace(",", ""))
        if found and quarters:
            series["enrolled_dwellings"] = [found.get(q) for q in quarters]

    return {"quarters": quarters, "series": series}


SA4_SHEETS = {
    "build_types": "Table P.4",
    "categories": "Table P.5",
    "max_residents": "Table P.6",
    "newbuild_places": "Table P.7",
    "pipeline_categories": "Table P.8",
    "demand_status": "Table P.9",
    "demand_categories": "Table P.10",
    "newbuild_detail": "Table P.11",
    "existing_legacy_detail": "Table P.12",
    "pipeline_detail": "Table P.16",
}

SA3_SHEETS = {
    "build_types": "Table P.13",
    "categories": "Table P.14",
    "max_residents": "Table P.15",
    "demand_status": "Table P.17",
    "demand_categories": "Table P.18",
}

# Row labels that carry data but are not places: the national total and the
# participants whose state could not be determined.
NATIONAL_ROW = "Total"
UNKNOWN_STATE_ROW = "Missing"


def row_values(table, key):
    return table["rows"].get(key, {}).get("values", {})


def row_flags(table, key):
    return table["rows"].get(key, {}).get("flags", {})


def find_column(columns, *needles):
    """First column whose name contains all the needles, case-insensitively."""
    for column in columns:
        lowered = column.lower()
        if all(n in lowered for n in needles):
            return column
    return None


def classify_rows(tables):
    """Split every row label in a level's tables into national / state / region.

    The workbook publishes state subtotals and a national total as ordinary
    rows, and they reconcile exactly with the sum of their children, so they are
    read directly rather than re-aggregated -- which also avoids summing
    suppressed cells as though they were zero.
    """
    labels = []
    for table in tables.values():
        for label in table["rows"]:
            if label not in labels:
                labels.append(label)

    states, regions = [], []
    for label in labels:
        if label in (NATIONAL_ROW, UNKNOWN_STATE_ROW):
            continue
        if " - " in label:
            state, name = label.split(" - ", 1)
            regions.append((label, state, name))
        else:
            states.append(label)
    return states, regions


def build_record(tables, key, *, level, name, state, parent, node_id, derive_places):
    """One geography's full profile, in the same shape at every level."""
    categories = tables["categories"]
    demand = tables["demand_categories"]
    status = tables["demand_status"]

    stock = row_values(categories, key)
    stock_flags = row_flags(categories, key)
    need = row_values(demand, key)
    need_flags = row_flags(demand, key)

    # New-build places are published outright in P.7, so use the NDIA's own
    # figure there and reserve the derivation for existing/legacy stock and the
    # pipeline, where no places figure is published at all.
    newbuild_places, existing_places, pipeline_places = {}, {}, {}
    pipeline_stock = {}
    if derive_places:
        newbuild_places = {c: v for c, v in
                           row_values(tables["newbuild_places"], key).items()
                           if c != "Total" and v is not None}
        existing_places = places_by_category(
            row_values(tables["existing_legacy_detail"], key))
        pipeline_places = places_by_category(row_values(tables["pipeline_detail"], key))
        pipeline_stock = row_values(tables["pipeline_categories"], key)

    cells = {}
    for category in DESIGN_CATEGORIES:
        new_pl = newbuild_places.get(category, 0.0)
        old_pl = existing_places.get(category, 0.0)
        places = (new_pl + old_pl) if derive_places else None
        participants = need.get(category)
        cells[category] = {
            "enrolled_dwellings": stock.get(category),
            "enrolled_dwellings_flag": stock_flags.get(category),
            "enrolled_places": places,
            "newbuild_places": new_pl if derive_places else None,
            "existing_legacy_places": old_pl if derive_places else None,
            "pipeline_dwellings": pipeline_stock.get(category),
            "pipeline_places": pipeline_places.get(category) if derive_places else None,
            "participants_with_need": participants,
            "participants_with_need_flag": need_flags.get(category),
            "places_per_participant": (
                round(places / participants, 3)
                if places is not None and participants not in (None, 0) else None
            ),
        }

    in_use_col = find_column(status["columns"], "in use")
    eligible_col = find_column(status["columns"], "eligible")
    need_total_col = find_column(demand["columns"], "total")

    return {
        "id": node_id,
        "level": level,
        "name": name,
        "state": state,
        "parent": parent,
        "source_row": key,
        "has_places": derive_places,
        "categories": cells,
        "totals": {
            "enrolled_dwellings": stock.get("Total"),
            "participants_sda_in_use": row_values(status, key).get(in_use_col),
            "participants_eligible_not_using": row_values(status, key).get(eligible_col),
            "participants_with_need": need.get(need_total_col),
            "need_without_category": need.get("Missing"),
            "pipeline_dwellings": (
                row_values(tables["pipeline_categories"], key).get("Total")
                if derive_places else None
            ),
        },
        "build_types": {c: v for c, v in row_values(tables["build_types"], key).items()},
        "max_residents": {c: v for c, v in row_values(tables["max_residents"], key).items()},
    }


def assemble(workbook):
    """Build one flat geography list spanning National, State, SA4 and SA3.

    Every node carries the same shape, so the interface renders one component at
    all four levels. Places are derivable only where the dwelling-form
    cross-tabs exist (P.11/P.12/P.16), which the NDIA publishes for SA4 and
    above but not for SA3.
    """
    sa4 = {key: read_table(workbook, name) for key, name in SA4_SHEETS.items()}
    sa3 = {key: read_table(workbook, name) for key, name in SA3_SHEETS.items()}

    sa4_states, sa4_regions = classify_rows(sa4)
    _, sa3_regions = classify_rows(sa3)

    nodes = [build_record(
        sa4, NATIONAL_ROW, level="National", name="Australia", state=None,
        parent=None, node_id="national", derive_places=True)]

    for state in sa4_states:
        nodes.append(build_record(
            sa4, state, level="State", name=state, state=state,
            parent="national", node_id=f"state:{state}", derive_places=True))

    for key, state, name in sa4_regions:
        nodes.append(build_record(
            sa4, key, level="SA4", name=name, state=state,
            parent=f"state:{state}", node_id=f"sa4:{key}", derive_places=True))

    for key, state, name in sa3_regions:
        nodes.append(build_record(
            sa3, key, level="SA3", name=name, state=state,
            parent=f"state:{state}", node_id=f"sa3:{key}", derive_places=False))

    return nodes, sa4


def calibrate_derivation(nodes, sa4_tables):
    """Measure the derivation against the one places figure the NDIA publishes.

    Places for existing/legacy stock (P.12) and the pipeline (P.16) have to be
    derived as dwellings x the resident count named in each column header. That
    arithmetic cannot be checked directly, because no published figure exists --
    but the same arithmetic over P.11 can be compared against published
    new-build places in P.7, which calibrates it.

    The two do not agree exactly: an enrolled dwelling's maximum residents can
    be lower than its dwelling type implies, so the derivation runs slightly
    high or low on individual rows. This reports how closely it tracks, so the
    derived figures carry a known error rather than an assumed one.
    """
    published = sa4_tables["newbuild_places"]
    checked = exact = 0
    worst = 0.0
    published_total = derived_total = 0.0
    for node in nodes:
        if not node["has_places"]:
            continue
        expected = row_values(published, node["source_row"])
        if not expected:
            continue
        derived = places_by_category(
            row_values(sa4_tables["newbuild_detail"], node["source_row"]))
        for category in DESIGN_CATEGORIES:
            want = expected.get(category)
            if want is None:
                continue
            got = derived.get(category, 0.0)
            checked += 1
            if abs(want - got) <= 0.5:
                exact += 1
            worst = max(worst, abs(want - got))
            if node["level"] == "SA4":
                published_total += want
                derived_total += got
    bias = (derived_total / published_total - 1) if published_total else None
    return {
        "values_checked": checked,
        "exact": exact,
        "exact_share": round(exact / checked, 4) if checked else None,
        "largest_difference_places": worst,
        "net_bias_on_sa4_totals": round(bias, 5) if bias is not None else None,
    }


def national_summary(nodes):
    national = next(n for n in nodes if n["id"] == "national")
    summary = {}
    for category in COMPARABLE_CATEGORIES:
        cell = national["categories"][category]
        summary[category] = {
            "enrolled_dwellings": cell["enrolled_dwellings"],
            "enrolled_places": cell["enrolled_places"],
            "participants_with_need": cell["participants_with_need"],
            "pipeline_places": cell["pipeline_places"],
            "places_per_participant": cell["places_per_participant"],
        }
    return summary


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("-o", "--out", type=Path, default=Path("docs/data"))
    args = parser.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)

    converted = args.out / "_transitional.xlsx"
    to_transitional(args.workbook, converted)
    workbook = openpyxl.load_workbook(converted, read_only=True, data_only=True)

    figure_prose = [str(cell) for row in workbook["Figure P.1"].iter_rows(values_only=True)
                    for cell in row if cell]

    nodes, sa4_tables = assemble(workbook)
    caption = sa4_tables["categories"]["caption"]
    as_at = re.search(r"as at (.+)$", caption)

    calibration = calibrate_derivation(nodes, sa4_tables)

    bundle = {
        "meta": {
            "source": args.workbook.name,
            "as_at": as_at.group(1).strip() if as_at else None,
            "design_categories": DESIGN_CATEGORIES,
            "comparable_categories": COMPARABLE_CATEGORIES,
            "notes": sa4_tables["demand_categories"]["notes"]
                     + sa4_tables["pipeline_detail"]["notes"],
            "derivation_calibration": calibration,
        },
        "national_trend": extract_national_trend(args.workbook, figure_prose),
        "national_summary": national_summary(nodes),
        "geographies": nodes,
    }

    target = args.out / "sda.json"
    target.write_text(json.dumps(bundle, separators=(",", ":")))
    converted.unlink()

    levels = {}
    for node in nodes:
        levels[node["level"]] = levels.get(node["level"], 0) + 1

    print(f"wrote {target}  ({target.stat().st_size / 1024:.0f} KB)")
    print(f"  as at: {bundle['meta']['as_at']}")
    print("  geographies: " + "  ".join(f"{k}={v}" for k, v in levels.items()))
    print(f"  quarters of national trend: {len(bundle['national_trend']['quarters'])}")
    print(f"  derivation calibration vs published P.7: "
          f"{calibration['exact']}/{calibration['values_checked']} exact "
          f"({calibration['exact_share']:.1%}), largest gap "
          f"{calibration['largest_difference_places']:.0f} places, net bias "
          f"{calibration['net_bias_on_sa4_totals']:+.3%}")
    for category, cell in bundle["national_summary"].items():
        print(f"  {category:<24} places={cell['enrolled_places']:>7.0f}"
              f"  need={cell['participants_with_need']:>6.0f}"
              f"  ratio={cell['places_per_participant']}")


if __name__ == "__main__":
    main()
